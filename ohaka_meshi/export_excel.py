from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

COLUMNS = ["日付", "料理名", "区分", "レストラン名", "場所", "参考URL", "感想", "写真", "登録日時"]
COLUMN_WIDTHS = [12, 24, 6, 20, 20, 40, 40, 6, 19]


def build_excel(entries):
    wb = Workbook()
    ws = wb.active
    ws.title = "お墓飯記録"

    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for e in entries:
        row_idx = ws.max_row + 1
        ws.append([
            e["eaten_date"],
            e["dish_name"],
            "外食" if e["is_eating_out"] else "内食",
            e["restaurant_name"] or "",
            e["location"] or "",
            e["reference_url"] or "",
            e["comment"] or "",
            "あり" if e["screenshot_filename"] else "なし",
            e["created_at"],
        ])
        if e["reference_url"]:
            cell = ws.cell(row=row_idx, column=6)
            cell.hyperlink = e["reference_url"]
            cell.style = "Hyperlink"

    for i, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
